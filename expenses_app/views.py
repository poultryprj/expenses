import traceback
from django.shortcuts import render
import json
from openpyxl import load_workbook
from django.http import JsonResponse
from rest_framework.decorators import api_view
from openpyxl.utils.dataframe import dataframe_to_rows
from rest_framework import status
import pandas as pd
from openpyxl import load_workbook, Workbook
from rest_framework.response import Response
from .models import ExcelData  # Import your ExcelData model
from datetime import datetime  # Import datetime module
import pytz  # Import pytz module
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
import openpyxl
from openpyxl.worksheet.views import SheetView, Selection
import os
from openpyxl.utils import get_column_letter


@api_view(['POST'])
def create_excel(request, sheet_name):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode("utf-8"))
            json_objects = data.get('json_objects')

            # Provide the full path to your Excel file
            file_path = 'main.xlsx'

            if not json_objects:
                return Response({"error": "JSON objects are required."}, status=status.HTTP_400_BAD_REQUEST)

            # Define the default columns outside of the if-else block
            default_columns = ['date', 'time', 'category', 'id',
                               'description', 'payment mode', 'bank', 'amount', 'complaint']

            # Load the Excel workbook using openpyxl
            workbook = load_workbook(filename=file_path)

            # Find the current sheet or create it if it doesn't exist
            if sheet_name not in workbook.sheetnames:
                if len(workbook.sheetnames) == 0:
                    # Create the first sheet
                    new_sheet_name = "Raw_data_01"
                    new_sheet = workbook.create_sheet(title=new_sheet_name)
                    new_sheet.append(['Raw Data Summary'])  # Add the title row
                    new_sheet.append(default_columns)  # Add the column names

                    # Set column widths for default columns
                    for column_letter, column_name in zip('ABCDEFGHI', default_columns):
                        column = new_sheet.column_dimensions[column_letter]
                        # Adjust the width as needed
                        column.width = 30 if column_name in [
                            'description', 'complaint'] else 15

                    # Set the alignment for the header row (centered)
                    header_row = new_sheet[2]
                    for cell in header_row:
                        cell.alignment = Alignment(horizontal='center')
                else:
                    # Find the next available sheet name with a sequential number
                    sheet_number = 1
                    while True:
                        new_sheet_name = f"Raw_data_{sheet_number:02d}"
                        if new_sheet_name not in workbook.sheetnames:
                            break
                        sheet_number += 1
                    new_sheet = workbook.create_sheet(title=new_sheet_name)
                    new_sheet.merge_cells('A1:I1')
                    title_cell = new_sheet.cell(row=1, column=1)
                    title_cell.value = 'RAW DATA'
                    title_cell.alignment = Alignment(horizontal='center')
                    new_sheet.append(default_columns)  # Add the column names

                    # Set column widths for default columns
                    for column_letter, column_name in zip('ABCDEFGHI', default_columns):
                        column = new_sheet.column_dimensions[column_letter]
                        # Adjust the width as needed
                        column.width = 40 if column_name in [
                            'description', 'complaint'] else 15

                    # Set the alignment for the header row (centered)
                    header_row = new_sheet[2]
                    for cell in header_row:
                        cell.alignment = Alignment(horizontal='center')

                sheet = new_sheet
            else:
                sheet = workbook[sheet_name]

            # Check if column names already exist in the sheet
            column_names = [cell.value for cell in sheet[2]]

            # Get the current date and time in UTC+05:30 (IST)
            # Define the IST timezone
            ist = pytz.timezone('Asia/Kolkata')

            # Get the current date and time in the IST timezone
            current_datetime = datetime.now(ist)

            # Format current_date as "dd/mm/yyyy" and current_time as "HH:mm:ss"
            current_date = current_datetime.strftime(
                '%d-%m-%Y')  # Format date as dd/mm/yyyy
            current_time = current_datetime.strftime(
                '%H:%M:%S')  # Format time as HH:mm:ss

            # Loop through the JSON objects and add data to the Excel sheet

            for obj in json_objects:
                for value in obj.values():
                    if isinstance(value, str):
                        if obj.get('payment_mode').lower() == "cash":
                            if obj.get('bank'):
                                return JsonResponse({'error': 'Bank name not required..!!'}, status=status.HTTP_400_BAD_REQUEST)
                            else:
                                row = [current_date, current_time, obj.get('category'), obj.get('id', 0), obj.get(
                                    'description'), obj.get('payment_mode'), '', obj.get('amount', 0.0), obj.get('complaint')]
                                # Append data to the sheet
                                sheet.append(row)
                        elif obj.get('bank'):
                            row = [current_date, current_time, obj.get('category'), obj.get('id', 0), obj.get(
                                'description'), obj.get('payment_mode'), obj.get('bank'), obj.get('amount', 0.0), obj.get('complaint')]
                            # Append data to the sheet
                            sheet.append(row)
                        else:
                            return JsonResponse({'error': 'Bank name required..!!'}, status=status.HTTP_400_BAD_REQUEST)

            # Freeze headers
            sheet.freeze_panes = 'A3'

            for rawIndex in range(3, sheet.max_row + 1):
                for colIndex, column_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'], start=1):
                    cell_dimensions = sheet.column_dimensions[column_letter]
                    column = sheet.cell(row=rawIndex, column=colIndex)
                    column.alignment = Alignment(
                        horizontal='center', vertical='center')
                    if column_names[colIndex - 1] in ['description', 'complaint']:
                        # Add top and left alignment and text wrap
                        alignment = Alignment(
                            wrap_text=True, vertical='top', horizontal='left')
                        column.alignment = alignment  # Apply the alignment

            # Save the updated Excel file after appending data
            workbook.save(file_path)

            return Response({'message': 'Data appended successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            print(traceback.format_exc())
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

################################


@api_view(['POST'])
def create_daily_summary_sheet(request, sheet_name):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode("utf-8"))

            # Provide the full path to your Excel file
            file_path = 'main.xlsx'

            # Check if the file exists
            if not os.path.isfile(file_path):
                return JsonResponse({'error': f'File not found at path: {file_path}'}, status=status.HTTP_400_BAD_REQUEST)

            # Load the Excel workbook using openpyxl
            workbook = load_workbook(filename=file_path)

            if sheet_name in workbook.sheetnames:
                return JsonResponse({'error': f'Sheet "{sheet_name}" already exists'}, status=status.HTTP_400_BAD_REQUEST)

            # Create a new sheet with the provided sheet_name
            new_sheet = workbook.create_sheet(title=sheet_name)

            # # Define the default columns and add them to the A2 row
            # Define the product types and their respective column ranges
            product_types = {
                'ACCOUNT': (1, 5),
                'FARMER': (7, 7),
                'VEHICALS': (9, 9),
                'SHOP': (11, 11),
                'OTHER EXPENSE': (13, 13),
                'VOUCHERS': (15, 15)
                # Add more product types as needed
            }

            default_columns = [
                'date',
                'opening balance',
                'collection amount',
                'expenses',
                'closing balance',
                ''
            ]

            all_categories = ['farmer', 'vehical', 'shop', 'other', 'vouchers']

            for category in all_categories:
                if category:
                    default_columns.extend([
                        # 'id',
                        # 'mode',
                        'amount',
                        ''
                    ])

            # Merge cells and set titles for each product type dynamically
            for product_type, (start_col, end_col) in product_types.items():
                new_sheet.merge_cells(
                    start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                title_cell = new_sheet.cell(row=1, column=start_col)
                title_cell.value = product_type
                title_cell.alignment = Alignment(horizontal='center')

            for col_num, header in enumerate(default_columns, start=1):
                new_sheet.cell(row=2, column=col_num, value=header)

            # Set the alignment for the header row (centered)
            header_row = new_sheet[2]
            for cell in header_row:
                cell.alignment = Alignment(horizontal='center')

            # Set column widths for default columns
            for i, column_name in enumerate(default_columns):
                # +1 because columns are 1-indexed
                column_letter = get_column_letter(i + 1)
                column = new_sheet.column_dimensions[column_letter]
                # Adjust the width as needed
                # Minimum width of 12
                column.width = max(len(column_name) + 2, 14)

            # Define the financial year start and end dates
            financial_year_start = datetime(2023, 4, 1)
            financial_year_end = datetime(2024, 3, 31)

            # Iterate over each date within the financial year
            current_date = financial_year_start
            while current_date <= financial_year_end:
                # Create a new row for each date
                row = [current_date.strftime('%d-%m-%Y'), '', '', '', '', '', '', '',
                       '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
                new_sheet.append(row)

                # Move to the next date
                current_date += timedelta(days=1)
            workbook.save(file_path)

            coll_name = ['FARMER', 'VEHICALS',
                         'SHOP', 'OTHER EXPENSE', 'VOUCHERS']
            for row in range(3, 369):
                if 'FARMER' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "F*", Raw_data_01!A:A, $A{row}), "")'
                    new_sheet[f'G{row}'] = sum_of_amount

                if 'VEHICALS' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "V*", Raw_data_01!A:A, $A{row}), "")'
                    new_sheet[f'I{row}'] = sum_of_amount

                if 'SHOP' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "S*", Raw_data_01!A:A, $A{row}), "")'
                    new_sheet[f'K{row}'] = sum_of_amount

                if 'OTHER EXPENSE' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "O*", Raw_data_01!A:A, $A{row}), "")'
                    new_sheet[f'M{row}'] = sum_of_amount

                if 'VOUCHERS' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "W*", Raw_data_01!A:A, $A{row}), "")'
                    new_sheet[f'O{row}'] = sum_of_amount

            # cloasing balance
            for row in range(3, 369):
                closing_balance = f'=SUM(B{row},C{row},G{row},I{row},K{row},M{row},O{row}) - D{row}'
                new_sheet[f'E{row}'] = closing_balance

            # Add the formula to the "B" column (opening_balance column) from B4 to B368
            for row in range(4, 369):
                formula = f'=IF(E{row - 1}<>0, E{row - 1}, IFERROR(INDEX(E3:E${row - 1}, MATCH(1, E3:E${row - 1}<>0, 0)), LOOKUP(2, 1/(E3:E${row - 1}<>0), E3:E${row - 1})))'
                new_sheet[f'B{row}'] = formula
            # Define a function to convert Excel column letters to column index

            def col_letter_to_index(col_letter):
                result = 0
                for letter in col_letter:
                    result = result * 26 + (ord(letter) - ord('A') + 1)
                return result

            # Format the columns
            columns_to_format = ['B', 'C', 'D', 'E', 'G', 'I', 'K', 'M', 'O']

            for col_letter in columns_to_format:
                col_index = col_letter_to_index(col_letter)
                # Format the columns to display two decimal places
                for row in new_sheet.iter_rows(min_row=3, max_row=369, min_col=col_index, max_col=col_index):
                    for cell in row:
                        cell.number_format = '0.00'

            # Freeze the top row (column names) when scrolling
            new_sheet.freeze_panes = "A3"

            # Save the updated Excel file again
            workbook.save(file_path)
            # ADD REQUIRED DEPENDENT SHEETS
            cash_payment(request, sheet_name="Cash_01")
            hdfc_payment(request, sheet_name="Hdfc_bank_01")
            idbi_payment(request, sheet_name="Idbi_bank_01")
            icici_payment(request, sheet_name="Icici_bank_01")

            return JsonResponse({'message': f'Successfully Created {sheet_name} with required sheets & Formulas'}, status=status.HTTP_200_OK)
        except FileNotFoundError as e:
            return JsonResponse({'error': 'File not found'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=status.HTTP_400_BAD_REQUEST)

# CASH PAYMENTS


def cash_payment(request, sheet_name):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode("utf-8"))

            # Provide the full path to your Excel file
            file_path = 'main.xlsx'

            # Check if the file exists
            if not os.path.isfile(file_path):
                return JsonResponse({'error': f'File not found at path: {file_path}'}, status=status.HTTP_400_BAD_REQUEST)

            # Load the Excel workbook using openpyxl
            workbook = load_workbook(filename=file_path)

            if sheet_name in workbook.sheetnames:
                return JsonResponse({'error': f'Sheet "{sheet_name}" already exists'}, status=status.HTTP_400_BAD_REQUEST)

            # Create a new sheet with the provided sheet_name
            new_sheet = workbook.create_sheet(title=sheet_name)

            # # Define the default columns and add them to the A2 row
            # Define the product types and their respective column ranges
            product_types = {
                'CASH ACCOUNT': (1, 5),
                'FARMER': (7, 7),
                'VEHICALS': (9, 9),
                'SHOP': (11, 11),
                'OTHER EXPENSE': (13, 13),
                'VOUCHERS': (15, 15)
                # Add more product types as needed
            }

            default_columns = [
                'date',
                'opening balance',
                'collection amount',
                'expenses',
                'closing balance',
                ''
            ]

            all_categories = ['farmer', 'vehical', 'shop', 'other', 'vouchers']

            for category in all_categories:
                if category:
                    default_columns.extend([
                        # 'id',
                        # 'mode',
                        'cash amount',
                        ''
                    ])

            # Merge cells and set titles for each product type dynamically
            for product_type, (start_col, end_col) in product_types.items():
                new_sheet.merge_cells(
                    start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                title_cell = new_sheet.cell(row=1, column=start_col)
                title_cell.value = product_type
                title_cell.alignment = Alignment(horizontal='center')

            for col_num, header in enumerate(default_columns, start=1):
                new_sheet.cell(row=2, column=col_num, value=header)

            # Set the alignment for the header row (centered)
            header_row = new_sheet[2]
            for cell in header_row:
                cell.alignment = Alignment(horizontal='center')

            # Set column widths for default columns
            for i, column_name in enumerate(default_columns):
                # +1 because columns are 1-indexed
                column_letter = get_column_letter(i + 1)
                column = new_sheet.column_dimensions[column_letter]
                # Adjust the width as needed
                # Minimum width of 12
                column.width = max(len(column_name) + 2, 14)

            # Define the financial year start and end dates
            financial_year_start = datetime(2023, 4, 1)
            financial_year_end = datetime(2024, 3, 31)

            # Iterate over each date within the financial year
            current_date = financial_year_start
            while current_date <= financial_year_end:
                # Create a new row for each date
                row = [current_date.strftime('%d-%m-%Y'), '', '', '', '', '', '', '',
                       '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
                new_sheet.append(row)

                # Move to the next date
                current_date += timedelta(days=1)
            workbook.save(file_path)

            coll_name = ['FARMER', 'VEHICALS',
                         'SHOP', 'OTHER EXPENSE', 'VOUCHERS']
            for row in range(3, 369):
                if 'FARMER' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "F*", Raw_data_01!A:A, $A{row}, Raw_data_01!F:F, "cash"), "")'
                    new_sheet[f'G{row}'] = sum_of_amount

                if 'VEHICALS' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "V*", Raw_data_01!A:A, $A{row}, Raw_data_01!F:F, "cash"), "")'
                    new_sheet[f'I{row}'] = sum_of_amount

                if 'SHOP' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "S*", Raw_data_01!A:A, $A{row}, Raw_data_01!F:F, "cash"), "")'
                    new_sheet[f'K{row}'] = sum_of_amount

                if 'OTHER EXPENSE' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "O*", Raw_data_01!A:A, $A{row}, Raw_data_01!F:F, "cash"), "")'
                    new_sheet[f'M{row}'] = sum_of_amount

                if 'VOUCHERS' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "W*", Raw_data_01!A:A, $A{row}, Raw_data_01!F:F, "cash"), "")'
                    new_sheet[f'O{row}'] = sum_of_amount

            # cloasing balance
            for row in range(3, 369):
                closing_balance = f'=SUM(B{row},C{row},G{row},I{row},K{row},M{row},O{row}) - D{row}'
                new_sheet[f'E{row}'] = closing_balance

            # Add the formula to the "B" column (opening_balance column) from B4 to B368
            for row in range(4, 369):
                formula = f'=IF(E{row - 1}<>0, E{row - 1}, IFERROR(INDEX(E3:E${row - 1}, MATCH(1, E3:E${row - 1}<>0, 0)), LOOKUP(2, 1/(E3:E${row - 1}<>0), E3:E${row - 1})))'
                new_sheet[f'B{row}'] = formula
            # Define a function to convert Excel column letters to column index

            def col_letter_to_index(col_letter):
                result = 0
                for letter in col_letter:
                    result = result * 26 + (ord(letter) - ord('A') + 1)
                return result

            # Format the columns
            columns_to_format = ['B', 'C', 'D', 'E', 'G', 'I', 'K', 'M', 'O']

            for col_letter in columns_to_format:
                col_index = col_letter_to_index(col_letter)
                # Format the columns to display two decimal places
                for row in new_sheet.iter_rows(min_row=3, max_row=369, min_col=col_index, max_col=col_index):
                    for cell in row:
                        cell.number_format = '0.00'

            # Freeze the top row (column names) when scrolling
            new_sheet.freeze_panes = "A3"

            # Save the updated Excel file again
            workbook.save(file_path)

            return JsonResponse({'message': f'Successfully Created sheet with Data & Formulas'}, status=status.HTTP_200_OK)
        except FileNotFoundError as e:
            return JsonResponse({'error': 'File not found'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=status.HTTP_400_BAD_REQUEST)

# HDFC BANK


def hdfc_payment(request, sheet_name):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode("utf-8"))

            # Provide the full path to your Excel file
            file_path = 'main.xlsx'

            # Check if the file exists
            if not os.path.isfile(file_path):
                return JsonResponse({'error': f'File not found at path: {file_path}'}, status=status.HTTP_400_BAD_REQUEST)

            # Load the Excel workbook using openpyxl
            workbook = load_workbook(filename=file_path)

            if sheet_name in workbook.sheetnames:
                return JsonResponse({'error': f'Sheet "{sheet_name}" already exists'}, status=status.HTTP_400_BAD_REQUEST)

            # Create a new sheet with the provided sheet_name
            new_sheet = workbook.create_sheet(title=sheet_name)

            # # Define the default columns and add them to the A2 row
            # Define the product types and their respective column ranges
            product_types = {
                'HDFC ACCOUNT': (1, 5),
                'FARMER': (7, 7),
                'VEHICALS': (9, 9),
                'SHOP': (11, 11),
                'OTHER EXPENSE': (13, 13),
                'VOUCHERS': (15, 15)
                # Add more product types as needed
            }

            default_columns = [
                'date',
                'opening balance',
                'collection amount',
                'expenses',
                'closing balance',
                ''
            ]

            all_categories = ['farmer', 'vehical', 'shop', 'other', 'vouchers']

            for category in all_categories:
                if category:
                    default_columns.extend([
                        # 'id',
                        # 'mode',
                        'HDFC amount',
                        ''
                    ])

            # Merge cells and set titles for each product type dynamically
            for product_type, (start_col, end_col) in product_types.items():
                new_sheet.merge_cells(
                    start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                title_cell = new_sheet.cell(row=1, column=start_col)
                title_cell.value = product_type
                title_cell.alignment = Alignment(horizontal='center')

            for col_num, header in enumerate(default_columns, start=1):
                new_sheet.cell(row=2, column=col_num, value=header)

            # Set the alignment for the header row (centered)
            header_row = new_sheet[2]
            for cell in header_row:
                cell.alignment = Alignment(horizontal='center')

            # Set column widths for default columns
            for i, column_name in enumerate(default_columns):
                # +1 because columns are 1-indexed
                column_letter = get_column_letter(i + 1)
                column = new_sheet.column_dimensions[column_letter]
                # Adjust the width as needed
                # Minimum width of 12
                column.width = max(len(column_name) + 2, 14)

            # Define the financial year start and end dates
            financial_year_start = datetime(2023, 4, 1)
            financial_year_end = datetime(2024, 3, 31)

            # Iterate over each date within the financial year
            current_date = financial_year_start
            while current_date <= financial_year_end:
                # Create a new row for each date
                row = [current_date.strftime('%d-%m-%Y'), '', '', '', '', '', '', '',
                       '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
                new_sheet.append(row)

                # Move to the next date
                current_date += timedelta(days=1)
            workbook.save(file_path)

            coll_name = ['FARMER', 'VEHICALS',
                         'SHOP', 'OTHER EXPENSE', 'VOUCHERS']
            for row in range(3, 369):
                if 'FARMER' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "F*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "hdfc"), "")'
                    new_sheet[f'G{row}'] = sum_of_amount

                if 'VEHICALS' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "V*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "hdfc"), "")'
                    new_sheet[f'I{row}'] = sum_of_amount

                if 'SHOP' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "S*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "hdfc"), "")'
                    new_sheet[f'K{row}'] = sum_of_amount

                if 'OTHER EXPENSE' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "O*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "hdfc"), "")'
                    new_sheet[f'M{row}'] = sum_of_amount

                if 'VOUCHERS' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "W*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "hdfc"), "")'
                    new_sheet[f'O{row}'] = sum_of_amount

            # cloasing balance
            for row in range(3, 369):
                closing_balance = f'=SUM(B{row},C{row},G{row},I{row},K{row},M{row},O{row}) - D{row}'
                new_sheet[f'E{row}'] = closing_balance

            # Add the formula to the "B" column (opening_balance column) from B4 to B368
            for row in range(4, 369):
                formula = f'=IF(E{row - 1}<>0, E{row - 1}, IFERROR(INDEX(E3:E${row - 1}, MATCH(1, E3:E${row - 1}<>0, 0)), LOOKUP(2, 1/(E3:E${row - 1}<>0), E3:E${row - 1})))'
                new_sheet[f'B{row}'] = formula
            # Define a function to convert Excel column letters to column index

            def col_letter_to_index(col_letter):
                result = 0
                for letter in col_letter:
                    result = result * 26 + (ord(letter) - ord('A') + 1)
                return result

            # Format the columns
            columns_to_format = ['B', 'C', 'D', 'E', 'G', 'I', 'K', 'M', 'O']

            for col_letter in columns_to_format:
                col_index = col_letter_to_index(col_letter)
                # Format the columns to display two decimal places
                for row in new_sheet.iter_rows(min_row=3, max_row=369, min_col=col_index, max_col=col_index):
                    for cell in row:
                        cell.number_format = '0.00'

            # Freeze the top row (column names) when scrolling
            new_sheet.freeze_panes = "A3"

            # Save the updated Excel file again
            workbook.save(file_path)

            return JsonResponse({'message': f'Successfully Created sheet with Data & Formulas'}, status=status.HTTP_200_OK)
        except FileNotFoundError as e:
            return JsonResponse({'error': 'File not found'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=status.HTTP_400_BAD_REQUEST)

# IDBI BANK


def idbi_payment(request, sheet_name):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode("utf-8"))

            # Provide the full path to your Excel file
            file_path = 'main.xlsx'

            # Check if the file exists
            if not os.path.isfile(file_path):
                return JsonResponse({'error': f'File not found at path: {file_path}'}, status=status.HTTP_400_BAD_REQUEST)

            # Load the Excel workbook using openpyxl
            workbook = load_workbook(filename=file_path)

            if sheet_name in workbook.sheetnames:
                return JsonResponse({'error': f'Sheet "{sheet_name}" already exists'}, status=status.HTTP_400_BAD_REQUEST)

            # Create a new sheet with the provided sheet_name
            new_sheet = workbook.create_sheet(title=sheet_name)

            # # Define the default columns and add them to the A2 row
            # Define the product types and their respective column ranges
            product_types = {
                'IDBI ACCOUNT': (1, 5),
                'FARMER': (7, 7),
                'VEHICALS': (9, 9),
                'SHOP': (11, 11),
                'OTHER EXPENSE': (13, 13),
                'VOUCHERS': (15, 15)
                # Add more product types as needed
            }

            default_columns = [
                'date',
                'opening balance',
                'collection amount',
                'expenses',
                'closing balance',
                ''
            ]

            all_categories = ['farmer', 'vehical', 'shop', 'other', 'vouchers']

            for category in all_categories:
                if category:
                    default_columns.extend([
                        # 'id',
                        # 'mode',
                        'IDBI amount',
                        ''
                    ])

            # Merge cells and set titles for each product type dynamically
            for product_type, (start_col, end_col) in product_types.items():
                new_sheet.merge_cells(
                    start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                title_cell = new_sheet.cell(row=1, column=start_col)
                title_cell.value = product_type
                title_cell.alignment = Alignment(horizontal='center')

            for col_num, header in enumerate(default_columns, start=1):
                new_sheet.cell(row=2, column=col_num, value=header)

            # Set the alignment for the header row (centered)
            header_row = new_sheet[2]
            for cell in header_row:
                cell.alignment = Alignment(horizontal='center')

            # Set column widths for default columns
            for i, column_name in enumerate(default_columns):
                # +1 because columns are 1-indexed
                column_letter = get_column_letter(i + 1)
                column = new_sheet.column_dimensions[column_letter]
                # Adjust the width as needed
                # Minimum width of 12
                column.width = max(len(column_name) + 2, 14)

            # Define the financial year start and end dates
            financial_year_start = datetime(2023, 4, 1)
            financial_year_end = datetime(2024, 3, 31)

            # Iterate over each date within the financial year
            current_date = financial_year_start
            while current_date <= financial_year_end:
                # Create a new row for each date
                row = [current_date.strftime('%d-%m-%Y'), '', '', '', '', '', '', '',
                       '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
                new_sheet.append(row)

                # Move to the next date
                current_date += timedelta(days=1)
            workbook.save(file_path)

            coll_name = ['FARMER', 'VEHICALS',
                         'SHOP', 'OTHER EXPENSE', 'VOUCHERS']
            for row in range(3, 369):
                if 'FARMER' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "F*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "idbi"), "")'
                    new_sheet[f'G{row}'] = sum_of_amount

                if 'VEHICALS' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "V*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "idbi"), "")'
                    new_sheet[f'I{row}'] = sum_of_amount

                if 'SHOP' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "S*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "idbi"), "")'
                    new_sheet[f'K{row}'] = sum_of_amount

                if 'OTHER EXPENSE' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "O*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "idbi"), "")'
                    new_sheet[f'M{row}'] = sum_of_amount

                if 'VOUCHERS' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "W*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "idbi"), "")'
                    new_sheet[f'O{row}'] = sum_of_amount

            # cloasing balance
            for row in range(3, 369):
                closing_balance = f'=SUM(B{row},C{row},G{row},I{row},K{row},M{row},O{row}) - D{row}'
                new_sheet[f'E{row}'] = closing_balance

            # Add the formula to the "B" column (opening_balance column) from B4 to B368
            for row in range(4, 369):
                formula = f'=IF(E{row - 1}<>0, E{row - 1}, IFERROR(INDEX(E3:E${row - 1}, MATCH(1, E3:E${row - 1}<>0, 0)), LOOKUP(2, 1/(E3:E${row - 1}<>0), E3:E${row - 1})))'
                new_sheet[f'B{row}'] = formula
            # Define a function to convert Excel column letters to column index

            def col_letter_to_index(col_letter):
                result = 0
                for letter in col_letter:
                    result = result * 26 + (ord(letter) - ord('A') + 1)
                return result

            # Format the columns
            columns_to_format = ['B', 'C', 'D', 'E', 'G', 'I', 'K', 'M', 'O']

            for col_letter in columns_to_format:
                col_index = col_letter_to_index(col_letter)
                # Format the columns to display two decimal places
                for row in new_sheet.iter_rows(min_row=3, max_row=369, min_col=col_index, max_col=col_index):
                    for cell in row:
                        cell.number_format = '0.00'

            # Freeze the top row (column names) when scrolling
            new_sheet.freeze_panes = "A3"

            # Save the updated Excel file again
            workbook.save(file_path)

            return JsonResponse({'message': f'Successfully Created sheet with Data & Formulas'}, status=status.HTTP_200_OK)
        except FileNotFoundError as e:
            return JsonResponse({'error': 'File not found'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=status.HTTP_400_BAD_REQUEST)


# ICICI BANK PAYMENT

def icici_payment(request, sheet_name):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode("utf-8"))

            # Provide the full path to your Excel file
            file_path = 'main.xlsx'

            # Check if the file exists
            if not os.path.isfile(file_path):
                return JsonResponse({'error': f'File not found at path: {file_path}'}, status=status.HTTP_400_BAD_REQUEST)

            # Load the Excel workbook using openpyxl
            workbook = load_workbook(filename=file_path)

            if sheet_name in workbook.sheetnames:
                return JsonResponse({'error': f'Sheet "{sheet_name}" already exists'}, status=status.HTTP_400_BAD_REQUEST)

            # Create a new sheet with the provided sheet_name
            new_sheet = workbook.create_sheet(title=sheet_name)

            # # Define the default columns and add them to the A2 row
            # Define the product types and their respective column ranges
            product_types = {
                'ICICI ACCOUNT': (1, 5),
                'FARMER': (7, 7),
                'VEHICALS': (9, 9),
                'SHOP': (11, 11),
                'OTHER EXPENSE': (13, 13),
                'VOUCHERS': (15, 15)
                # Add more product types as needed
            }

            default_columns = [
                'date',
                'opening balance',
                'collection amount',
                'expenses',
                'closing balance',
                ''
            ]

            all_categories = ['farmer', 'vehical', 'shop', 'other', 'vouchers']

            for category in all_categories:
                if category:
                    default_columns.extend([
                        # 'id',
                        # 'mode',
                        'ICICI amount',
                        ''
                    ])

            # Merge cells and set titles for each product type dynamically
            for product_type, (start_col, end_col) in product_types.items():
                new_sheet.merge_cells(
                    start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                title_cell = new_sheet.cell(row=1, column=start_col)
                title_cell.value = product_type
                title_cell.alignment = Alignment(horizontal='center')

            for col_num, header in enumerate(default_columns, start=1):
                new_sheet.cell(row=2, column=col_num, value=header)

            # Set the alignment for the header row (centered)
            header_row = new_sheet[2]
            for cell in header_row:
                cell.alignment = Alignment(horizontal='center')

            # Set column widths for default columns
            for i, column_name in enumerate(default_columns):
                # +1 because columns are 1-indexed
                column_letter = get_column_letter(i + 1)
                column = new_sheet.column_dimensions[column_letter]
                # Adjust the width as needed
                # Minimum width of 12
                column.width = max(len(column_name) + 2, 14)

            # Define the financial year start and end dates
            financial_year_start = datetime(2023, 4, 1)
            financial_year_end = datetime(2024, 3, 31)

            # Iterate over each date within the financial year
            current_date = financial_year_start
            while current_date <= financial_year_end:
                # Create a new row for each date
                row = [current_date.strftime('%d-%m-%Y'), '', '', '', '', '', '', '',
                       '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
                new_sheet.append(row)

                # Move to the next date
                current_date += timedelta(days=1)
            workbook.save(file_path)

            coll_name = ['FARMER', 'VEHICALS',
                         'SHOP', 'OTHER EXPENSE', 'VOUCHERS']
            for row in range(3, 369):
                if 'FARMER' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "F*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "icici"), "")'
                    new_sheet[f'G{row}'] = sum_of_amount

                if 'VEHICALS' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "V*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "icici"), "")'
                    new_sheet[f'I{row}'] = sum_of_amount

                if 'SHOP' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "S*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "icici"), "")'
                    new_sheet[f'K{row}'] = sum_of_amount

                if 'OTHER EXPENSE' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "O*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "icici"), "")'
                    new_sheet[f'M{row}'] = sum_of_amount

                if 'VOUCHERS' in coll_name:
                    sum_of_amount = f'=IF($A{row}<>"", SUMIFS(Raw_data_01!H:H, Raw_data_01!C:C, "W*", Raw_data_01!A:A, $A{row}, Raw_data_01!G:G, "icici"), "")'
                    new_sheet[f'O{row}'] = sum_of_amount

            # cloasing balance
            for row in range(3, 369):
                closing_balance = f'=SUM(B{row},C{row},G{row},I{row},K{row},M{row},O{row}) - D{row}'
                new_sheet[f'E{row}'] = closing_balance

            # Add the formula to the "B" column (opening_balance column) from B4 to B368
            for row in range(4, 369):
                formula = f'=IF(E{row - 1}<>0, E{row - 1}, IFERROR(INDEX(E3:E${row - 1}, MATCH(1, E3:E${row - 1}<>0, 0)), LOOKUP(2, 1/(E3:E${row - 1}<>0), E3:E${row - 1})))'
                new_sheet[f'B{row}'] = formula
            # Define a function to convert Excel column letters to column index

            def col_letter_to_index(col_letter):
                result = 0
                for letter in col_letter:
                    result = result * 26 + (ord(letter) - ord('A') + 1)
                return result

            # Format the columns
            columns_to_format = ['B', 'C', 'D', 'E', 'G', 'I', 'K', 'M', 'O']

            for col_letter in columns_to_format:
                col_index = col_letter_to_index(col_letter)
                # Format the columns to display two decimal places
                for row in new_sheet.iter_rows(min_row=3, max_row=369, min_col=col_index, max_col=col_index):
                    for cell in row:
                        cell.number_format = '0.00'

            # Freeze the top row (column names) when scrolling
            new_sheet.freeze_panes = "A3"

            # Save the updated Excel file again
            workbook.save(file_path)

            return JsonResponse({'message': f'Successfully Created sheet with Data & Formulas'}, status=status.HTTP_200_OK)
        except FileNotFoundError as e:
            return JsonResponse({'error': 'File not found'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=status.HTTP_400_BAD_REQUEST)
